"""
AUCTUS Capital Partners AG — Deterministic DCF Engine

Computes Unlevered Free Cash Flow (UFCF), discounts at WACC using Gordon Growth
Model terminal value, and writes JSON + CSV output artifacts.

This script is the ONLY authoritative source of DCF numbers. The agent never
computes financial values inline — it invokes this script and reads the output.

Usage:
    python scripts/dcf_engine.py \
        --input data/inputs/company_financials.csv \
        --projections data/inputs/company_projections_approved.csv \
        --wacc 0.12 \
        --terminal-growth 0.025 \
        --projection-years 5 \
        --output-dir outputs/dcf_models/ \
        --company-name "Muster GmbH"

Exit codes:
    0  success — all output files written
    1  input validation failure
    2  computation error
    3  output write failure
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("logs/dcf_engine.log", mode="a"),
        logging.StreamHandler(sys.stderr),
    ],
)
logger = logging.getLogger(__name__)

REQUIRED_COLUMNS = {"year", "revenue", "ebitda", "d_and_a", "capex", "nwc_change", "tax_rate"}


# ── Input Loading ─────────────────────────────────────────────────────────────

def load_financials(path: Path) -> pd.DataFrame:
    """Load historical financial CSV and validate schema."""
    df = pd.read_csv(path)
    df.columns = df.columns.str.strip().str.lower()
    missing = REQUIRED_COLUMNS - set(df.columns)
    if missing:
        raise ValueError(f"Input CSV missing required columns: {missing}")
    df = df.sort_values("year").reset_index(drop=True)
    if len(df) < 3:
        logger.warning("Fewer than 3 historical years — projection confidence is LOW")
    return df


def load_projections(path: Optional[Path], hist: pd.DataFrame) -> pd.DataFrame:
    """Load approved projection CSV, or auto-derive from historicals."""
    if path and path.exists():
        proj = pd.read_csv(path)
        proj.columns = proj.columns.str.strip().str.lower()
        missing = REQUIRED_COLUMNS - set(proj.columns)
        if missing:
            raise ValueError(f"Projections CSV missing columns: {missing}")
        return proj.sort_values("year").reset_index(drop=True)

    logger.info("No projections file found — deriving from 3-year historical trends")
    revenue_cagr = _compute_cagr(hist["revenue"].iloc[0], hist["revenue"].iloc[-1], len(hist) - 1)
    avg_ebitda_margin = (hist["ebitda"] / hist["revenue"]).mean()
    avg_da_pct = (hist["d_and_a"] / hist["revenue"]).mean()
    avg_capex_pct = (hist["capex"] / hist["revenue"]).mean()
    avg_nwc_pct = (hist["nwc_change"] / hist["revenue"]).abs().mean()
    tax_rate = hist["tax_rate"].iloc[-1]

    base_revenue = hist["revenue"].iloc[-1]
    base_year = int(hist["year"].iloc[-1])
    rows = []
    for i in range(1, 6):
        rev = base_revenue * ((1 + revenue_cagr) ** i)
        ebitda = rev * avg_ebitda_margin
        da = rev * avg_da_pct
        capex = rev * avg_capex_pct
        nwc = rev * avg_nwc_pct
        rows.append({
            "year": base_year + i,
            "revenue": round(rev, 4),
            "ebitda": round(ebitda, 4),
            "d_and_a": round(da, 4),
            "capex": round(capex, 4),
            "nwc_change": round(nwc, 4),
            "tax_rate": tax_rate,
        })
    return pd.DataFrame(rows)


# ── Core Computation ──────────────────────────────────────────────────────────

def _compute_cagr(start: float, end: float, years: int) -> float:
    if years <= 0 or start <= 0:
        return 0.03
    return float((end / start) ** (1.0 / years) - 1.0)


def compute_ufcf(df: pd.DataFrame) -> pd.Series:
    """
    Unlevered Free Cash Flow = EBIT × (1 − tax) + D&A − CapEx − ΔNWC

    EBIT = EBITDA − D&A
    """
    ebit = df["ebitda"] - df["d_and_a"]
    nopat = ebit * (1.0 - df["tax_rate"])
    ufcf = nopat + df["d_and_a"] - df["capex"] - df["nwc_change"]
    return ufcf.round(6)


def discount_cashflows(fcf: pd.Series, wacc: float) -> float:
    """Present value of a series of annual FCFs discounted at WACC."""
    pv = sum(cf / ((1.0 + wacc) ** (t + 1)) for t, cf in enumerate(fcf))
    return float(round(pv, 6))


def terminal_value_gordon(final_fcf: float, wacc: float, tgr: float) -> float:
    """
    Gordon Growth Model: TV = FCF_n × (1 + g) / (WACC − g)
    Present-values the terminal value to t=0.
    """
    if wacc <= tgr:
        raise ValueError(
            f"WACC ({wacc:.4f}) must exceed terminal growth rate ({tgr:.4f}). "
            "Model is undefined when WACC ≤ TGR."
        )
    tv_at_n = final_fcf * (1.0 + tgr) / (wacc - tgr)
    n = 5
    tv_pv = tv_at_n / ((1.0 + wacc) ** n)
    return float(round(tv_pv, 6))


def enterprise_value(pv_fcf: float, tv_pv: float) -> float:
    return float(round(pv_fcf + tv_pv, 6))


# ── Orchestration ─────────────────────────────────────────────────────────────

def run_dcf(
    input_path: Path,
    projections_path: Optional[Path],
    wacc: float,
    tgr: float,
    years: int,
    output_dir: Path,
    company_name: str = "unknown",
) -> dict:
    hist = load_financials(input_path)
    proj = load_projections(projections_path, hist)
    proj = proj.head(years)

    ufcf = compute_ufcf(proj)
    pv_fcf = discount_cashflows(ufcf, wacc)
    tv_pv = terminal_value_gordon(float(ufcf.iloc[-1]), wacc, tgr)
    ev = enterprise_value(pv_fcf, tv_pv)

    tv_pct = round(tv_pv / ev, 6) if ev != 0 else float("nan")


    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    slug = company_name.lower().replace(" ", "_").replace("/", "_")

    # Write cashflows CSV
    cashflow_df = proj.copy()
    cashflow_df["ufcf"] = ufcf.values
    cashflow_df["pv_ufcf"] = [
        round(cf / ((1.0 + wacc) ** (t + 1)), 6)
        for t, cf in enumerate(ufcf)
    ]
    cf_path = output_dir / f"{slug}_{ts}_cashflows.xlsx"

    import openpyxl
    from openpyxl.styles import Font, PatternFill
    with pd.ExcelWriter(cf_path, engine="openpyxl") as _xw:
        cashflow_df.to_excel(_xw, index=False, sheet_name="Cash Flows")
        _ws = _xw.sheets["Cash Flows"]
        _navy = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        for cell in _ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
            cell.fill = _navy
        _col_fmts = {"ufcf": "€#,##0.00", "pv_ufcf": "€#,##0.00",
                     "revenue": "€#,##0.00", "ebitda": "€#,##0.00"}
        for ci, col_name in enumerate(cashflow_df.columns, start=1):
            _ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 16
            fmt = _col_fmts.get(col_name)
            if fmt:
                for ri in range(2, len(cashflow_df) + 2):
                    _ws.cell(row=ri, column=ci).number_format = fmt
        _ws.freeze_panes = "A2"

    # Write results JSON
    result: dict = {
        "company_name": company_name,
        "run_timestamp": ts,
        "wacc_used": wacc,
        "terminal_growth_rate_used": tgr,
        "projection_years": years,
        "pv_forecast_cashflows_eur_m": round(pv_fcf, 4),
        "terminal_value_pv_eur_m": round(tv_pv, 4),
        "enterprise_value_eur_m": round(ev, 4),
        "terminal_value_pct_of_ev": round(tv_pct, 4),
        "forecast_cash_flows": ufcf.round(4).tolist(),
        "sensitivity_inputs": {
            "terminal_fcf_eur_m": round(float(ufcf.iloc[-1]), 4),
            "base_wacc": wacc,
            "base_tgr": tgr,
        },
        "cashflows_path": str(cf_path),
    }
    json_path = output_dir / f"{slug}_{ts}_dcf_results.json"
    json_path.write_text(json.dumps(result, indent=2))

    logger.info(
        "DCF complete: EV=€%.2fm, TV%%=%.1f%%, WACC=%.2f%%, TGR=%.2f%%",
        ev, tv_pct * 100, wacc * 100, tgr * 100,
    )
    return result


# ── Entry Point ───────────────────────────────────────────────────────────────

def main() -> int:
    parser = argparse.ArgumentParser(description="AUCTUS DCF Engine")
    parser.add_argument("--input", type=Path, required=True, help="Historical financials CSV")
    parser.add_argument("--projections", type=Path, default=None, help="Approved projections CSV")
    parser.add_argument("--wacc", type=float, required=True, help="WACC as decimal (e.g. 0.12)")
    parser.add_argument("--terminal-growth", type=float, required=True, dest="tgr")
    parser.add_argument("--projection-years", type=int, default=5, dest="years")
    parser.add_argument("--output-dir", type=Path, required=True, dest="output_dir")
    parser.add_argument("--company-name", type=str, default="company", dest="company_name")
    args = parser.parse_args()

    try:
        args.output_dir.mkdir(parents=True, exist_ok=True)
        result = run_dcf(
            args.input,
            args.projections,
            args.wacc,
            args.tgr,
            args.years,
            args.output_dir,
            args.company_name,
        )
        print(json.dumps({"status": "success", "enterprise_value_eur_m": result["enterprise_value_eur_m"]}))
        return 0
    except ValueError as exc:
        logger.error("Input validation error: %s", exc)
        return 1
    except (ArithmeticError, ZeroDivisionError) as exc:
        logger.error("Computation error: %s", exc)
        return 2
    except OSError as exc:
        logger.error("Output write error: %s", exc)
        return 3


if __name__ == "__main__":
    sys.exit(main())
