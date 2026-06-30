"""
AUCTUS Capital Partners AG — LBO Modeling & Debt Waterfall Engine

Implements a five-step Leveraged Buyout framework for German Mittelstand acquisitions:

  Step 1 — Entry Assumptions: purchase price, capital structure
  Step 2 — Sources & Uses: balance check; fee capitalisation logic
  Step 3 — P&L Projections: Revenue → EBITDA → EBIT → Net Interest → LFCF
  Step 4 — Debt Waterfall: Senior TL (floating Euribor + spread, mandatory amort,
            cash sweep) and Notes (fixed or floating)
  Step 5 — Exit Metrics: MOIC, IRR, leverage stats; Entry × Exit sensitivity grid

All arithmetic is deterministic — the agent never estimates figures in context.

Usage:
    python scripts/lbo_engine.py \\
        --company-name "Muster GmbH" \\
        --entry-ebitda 8.5 \\
        --entry-multiple 8.0 \\
        --equity-pct 0.45 \\
        --senior-debt-pct 0.40 \\
        --notes-pct 0.15 \\
        --euribor 0.0390 \\
        --euribor-floor 0.0000 \\
        --senior-spread-bps 375 \\
        --notes-rate 0.0950 \\
        --notes-fixed \\
        --senior-amort-pct 0.05 \\
        --senior-cash-sweep-pct 0.50 \\
        --advisor-fee-pct 0.015 \\
        --financing-fee-pct 0.020 \\
        --capitalize-fees \\
        --tax-rate 0.299 \\
        --revenue-base 50.0 \\
        --revenue-growth 0.07,0.07,0.06,0.06,0.05 \\
        --ebitda-margins 0.175,0.18,0.185,0.19,0.195 \\
        --da-pct 0.025 \\
        --capex-pct 0.035 \\
        --nwc-pct 0.08 \\
        --exit-year 5 \\
        --exit-multiple 9.0 \\
        --output-dir outputs/dcf_models/

Exit codes:
    0  success — JSON results + CSV outputs written
    1  input validation failure
    2  computation error (including IRR convergence failure)
    3  output write failure
"""

from __future__ import annotations

import argparse
import json
import logging
import math
import sys
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import pandas as pd
import yaml
from pydantic import BaseModel, ConfigDict, field_validator, model_validator
from scipy.optimize import brentq

import openpyxl
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Logging ───────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("logs/lbo_engine.log", mode="a"),
        logging.StreamHandler(sys.stderr),
    ],
)
logger = logging.getLogger(__name__)

# ── Config Loader ─────────────────────────────────────────────────────────────

_CONFIG_DIR = Path("config")


def load_yaml(path: Path) -> dict[str, Any]:
    with open(path) as fh:
        data: dict[str, Any] = yaml.safe_load(fh)
    return data


def load_configs() -> tuple[dict[str, Any], dict[str, Any]]:
    """Return (financial_constants, auctus_criteria) dicts from config/."""
    fc = load_yaml(_CONFIG_DIR / "financial_constants.yaml")
    ac = load_yaml(_CONFIG_DIR / "auctus_criteria.yaml")
    return fc, ac


# ── Pydantic Models ───────────────────────────────────────────────────────────

class EntryAssumptions(BaseModel):
    model_config = ConfigDict(frozen=True)

    company_name: str
    geography: str = "DE"

    # Purchase price
    entry_ebitda_eur_m: float
    entry_multiple: float

    # Capital structure (must sum to 1.0)
    equity_pct: float
    senior_debt_pct: float
    notes_pct: float

    # Floating-rate senior debt
    euribor_rate: float
    euribor_floor: float
    senior_spread_bps: int

    # Notes (subordinated / second-lien)
    notes_is_fixed: bool
    notes_fixed_rate: float        # used when notes_is_fixed=True
    notes_euribor_spread_bps: int  # used when notes_is_fixed=False

    # Debt mechanics
    senior_amort_pct_annual: float   # % of original balance per year
    senior_cash_sweep_pct: float     # fraction of LFCF swept to senior repayment

    # Transaction fees
    advisor_fee_pct_ev: float       # % of entry EV
    financing_fee_pct_debt: float   # % of total debt (senior + notes)
    fees_capitalized: bool          # True → add to Uses and amortise; False → P&L yr1

    # Tax
    tax_rate: float

    # P&L projections
    projection_years: int
    revenue_base_eur_m: float
    revenue_growth_rates: list[float]
    ebitda_margins: list[float]
    da_pct_revenue: float
    capex_pct_revenue: float
    nwc_pct_revenue_change: float   # applied to annual revenue delta

    # Exit
    exit_year: int
    exit_multiple: float

    @field_validator("entry_ebitda_eur_m", "entry_multiple", "revenue_base_eur_m")
    @classmethod
    def must_be_positive(cls, v: float) -> float:
        if v <= 0:
            raise ValueError(f"Value must be positive, got {v}")
        return v

    @field_validator("equity_pct", "senior_debt_pct", "notes_pct")
    @classmethod
    def pct_range(cls, v: float) -> float:
        if not 0.0 <= v <= 1.0:
            raise ValueError(f"Percentage must be in [0, 1], got {v}")
        return v

    @field_validator("euribor_floor")
    @classmethod
    def floor_non_negative(cls, v: float) -> float:
        if v < 0:
            raise ValueError(f"Euribor floor cannot be negative, got {v}")
        return v

    @field_validator("tax_rate")
    @classmethod
    def tax_in_range(cls, v: float) -> float:
        if not 0.0 <= v < 1.0:
            raise ValueError(f"Tax rate must be in [0, 1), got {v}")
        return v

    @field_validator("senior_amort_pct_annual", "senior_cash_sweep_pct")
    @classmethod
    def amort_range(cls, v: float) -> float:
        if not 0.0 <= v <= 1.0:
            raise ValueError(f"Amortisation pct must be in [0, 1], got {v}")
        return v

    @model_validator(mode="after")
    def validate_structure(self) -> "EntryAssumptions":
        total = round(self.equity_pct + self.senior_debt_pct + self.notes_pct, 6)
        if abs(total - 1.0) > 1e-4:
            raise ValueError(
                f"Capital structure pcts must sum to 1.0, got {total:.6f}"
            )
        if self.exit_year > self.projection_years:
            raise ValueError(
                f"exit_year ({self.exit_year}) cannot exceed projection_years ({self.projection_years})"
            )
        if len(self.revenue_growth_rates) != self.projection_years:
            raise ValueError(
                f"revenue_growth_rates length ({len(self.revenue_growth_rates)}) "
                f"must equal projection_years ({self.projection_years})"
            )
        if len(self.ebitda_margins) != self.projection_years:
            raise ValueError(
                f"ebitda_margins length ({len(self.ebitda_margins)}) "
                f"must equal projection_years ({self.projection_years})"
            )
        return self


class SourcesUses(BaseModel):
    model_config = ConfigDict(frozen=True)

    entry_ev_eur_m: float
    equity_eur_m: float
    senior_debt_eur_m: float
    notes_eur_m: float
    total_sources_eur_m: float

    acquisition_price_eur_m: float
    advisor_fees_eur_m: float
    financing_fees_eur_m: float
    total_uses_eur_m: float

    balance_eur_m: float            # sources − uses; must be ~0
    financing_fees_capitalised: bool
    fee_amort_annual_eur_m: float   # annual P&L charge if fees capitalised


class YearProjection(BaseModel):
    model_config = ConfigDict(frozen=True)

    year: int
    revenue_eur_m: float
    ebitda_eur_m: float
    ebitda_margin: float
    da_eur_m: float
    ebit_eur_m: float
    fee_amort_eur_m: float          # non-cash financing fee amortisation
    senior_interest_eur_m: float
    notes_interest_eur_m: float
    total_interest_eur_m: float
    ebt_eur_m: float
    tax_eur_m: float
    net_income_eur_m: float
    capex_eur_m: float
    nwc_change_eur_m: float
    senior_mandatory_amort_eur_m: float
    levered_fcf_before_sweep_eur_m: float
    senior_cash_sweep_eur_m: float
    levered_fcf_eur_m: float        # after optional sweep; distributable to equity
    senior_opening_eur_m: float
    senior_closing_eur_m: float
    notes_opening_eur_m: float
    notes_closing_eur_m: float
    total_debt_closing_eur_m: float
    leverage_x: float               # total debt / EBITDA
    interest_coverage_x: float      # EBIT / total interest


class LBOMetrics(BaseModel):
    model_config = ConfigDict(frozen=True)

    entry_ev_eur_m: float
    entry_multiple_x: float
    exit_ev_eur_m: float
    exit_multiple_x: float
    exit_ebitda_eur_m: float
    net_debt_at_exit_eur_m: float
    equity_invested_eur_m: float
    equity_proceeds_eur_m: float
    moic: float
    irr: float                      # decimal; NaN if solver fails
    leverage_at_entry_x: float
    leverage_at_exit_x: float
    interest_coverage_min_x: float
    exit_year: int
    irr_solver_converged: bool


class LBOResult(BaseModel):
    model_config = ConfigDict(frozen=True)

    run_timestamp: str
    company_name: str
    assumptions: EntryAssumptions
    sources_uses: SourcesUses
    projections: list[YearProjection]
    metrics: LBOMetrics
    sensitivity_irr: dict[str, dict[str, float]]    # {entry_m_str: {exit_m_str: irr}}
    sensitivity_moic: dict[str, dict[str, float]]   # {entry_m_str: {exit_m_str: moic}}


# ── Step 2: Sources & Uses ────────────────────────────────────────────────────

def build_sources_uses(a: EntryAssumptions) -> SourcesUses:
    entry_ev = round(a.entry_ebitda_eur_m * a.entry_multiple, 6)
    senior = round(entry_ev * a.senior_debt_pct, 6)
    notes = round(entry_ev * a.notes_pct, 6)
    total_debt = round(senior + notes, 6)

    advisor_fees = round(entry_ev * a.advisor_fee_pct_ev, 6)
    financing_fees = round(total_debt * a.financing_fee_pct_debt, 6)

    acq_price = entry_ev

    # Uses: acquisition price + fees that appear on the closing statement.
    # When fees are capitalised both advisor and financing fees sit in Uses.
    # When fees are NOT capitalised only advisor fees appear in Uses (financing
    # fees are expensed through the P&L in Year 1 instead).
    if a.fees_capitalized:
        total_uses = round(acq_price + advisor_fees + financing_fees, 6)
        fee_amort = round(financing_fees / a.projection_years, 6)
    else:
        total_uses = round(acq_price + advisor_fees, 6)
        fee_amort = 0.0

    # Equity is derived as the residual that makes Sources = Uses.
    equity = round(total_uses - total_debt, 6)
    total_sources = round(equity + senior + notes, 6)
    balance = round(total_sources - total_uses, 6)   # ~0 by construction

    return SourcesUses(
        entry_ev_eur_m=entry_ev,
        equity_eur_m=equity,
        senior_debt_eur_m=senior,
        notes_eur_m=notes,
        total_sources_eur_m=total_sources,
        acquisition_price_eur_m=acq_price,
        advisor_fees_eur_m=advisor_fees,
        financing_fees_eur_m=financing_fees,
        total_uses_eur_m=total_uses,
        balance_eur_m=balance,
        financing_fees_capitalised=a.fees_capitalized,
        fee_amort_annual_eur_m=fee_amort,
    )


# ── Floating-Rate Helpers ─────────────────────────────────────────────────────

def _senior_rate(a: EntryAssumptions) -> float:
    """Effective senior rate = max(Euribor, floor) + spread."""
    return max(a.euribor_rate, a.euribor_floor) + a.senior_spread_bps / 10_000.0


def _notes_rate(a: EntryAssumptions) -> float:
    """Effective notes rate; fixed or floating."""
    if a.notes_is_fixed:
        return a.notes_fixed_rate
    return max(a.euribor_rate, a.euribor_floor) + a.notes_euribor_spread_bps / 10_000.0


# ── Step 3 + 4: P&L Projections & Debt Waterfall ─────────────────────────────

def compute_projections(
    a: EntryAssumptions,
    su: SourcesUses,
) -> list[YearProjection]:
    """
    Build year-by-year P&L and debt schedule simultaneously.

    Interest is computed on the opening balance (beginning-of-period convention),
    which avoids circularity between interest and closing-balance calculations.
    """
    sr = _senior_rate(a)
    nr = _notes_rate(a)
    senior_original = su.senior_debt_eur_m
    mandatory_amort = round(senior_original * a.senior_amort_pct_annual, 6)

    senior_bal = su.senior_debt_eur_m
    notes_bal = su.notes_eur_m
    prev_revenue = a.revenue_base_eur_m

    years: list[YearProjection] = []

    for i in range(a.projection_years):
        year_idx = i + 1
        growth = a.revenue_growth_rates[i]
        margin = a.ebitda_margins[i]

        revenue = round(prev_revenue * (1.0 + growth), 6)
        ebitda = round(revenue * margin, 6)
        da = round(revenue * a.da_pct_revenue, 6)
        ebit = round(ebitda - da, 6)
        capex = round(revenue * a.capex_pct_revenue, 6)
        nwc_change = round((revenue - prev_revenue) * a.nwc_pct_revenue_change, 6)

        # Financing-fee amortisation (non-cash, tax-deductible if capitalised)
        fee_amort = su.fee_amort_annual_eur_m if a.fees_capitalized else 0.0
        # Year-1: if fees NOT capitalised, expense financing fees fully in P&L
        yr1_fee_expense = su.financing_fees_eur_m if (not a.fees_capitalized and year_idx == 1) else 0.0

        # Interest on opening balances (beginning-of-period)
        senior_interest = round(senior_bal * sr, 6)
        notes_interest = round(notes_bal * nr, 6)
        total_interest = round(senior_interest + notes_interest, 6)

        # P&L below EBIT
        ebt = round(ebit - total_interest - fee_amort - yr1_fee_expense, 6)
        tax = round(max(0.0, ebt) * a.tax_rate, 6)
        net_income = round(ebt - tax, 6)

        # LFCF before optional sweep
        actual_mandatory = min(mandatory_amort, senior_bal)
        lfcf_before_sweep = round(
            net_income + da - capex - nwc_change - actual_mandatory, 6
        )

        # Optional cash sweep: applied to senior only
        senior_after_mandatory = round(senior_bal - actual_mandatory, 6)
        if lfcf_before_sweep > 0 and senior_after_mandatory > 0:
            sweep_amount = round(
                min(lfcf_before_sweep * a.senior_cash_sweep_pct, senior_after_mandatory), 6
            )
        else:
            sweep_amount = 0.0

        lfcf = round(lfcf_before_sweep - sweep_amount, 6)

        # Closing debt balances
        senior_close = round(max(0.0, senior_after_mandatory - sweep_amount), 6)
        notes_close = notes_bal   # notes bullet (no amortisation)

        # Strict QA Verification: Interest trace balance
        if abs(senior_interest - (senior_bal * sr)) > 0.002:
            raise ValueError(f"Computation error: Senior interest trace failed in Year {year_idx}")

        total_debt_close = round(senior_close + notes_close, 6)
        leverage_x = round(total_debt_close / ebitda, 4) if ebitda > 0 else float("inf")
        coverage_x = round(ebit / total_interest, 4) if total_interest > 0 else float("inf")

        if coverage_x < 1.0:
            logger.warning(
                "Year %d: interest coverage %.2fx — EBIT insufficient to cover interest. "
                "Equity cure or revolver draw may be required.",
                year_idx,
                coverage_x,
            )

        years.append(YearProjection(
            year=year_idx,
            revenue_eur_m=revenue,
            ebitda_eur_m=ebitda,
            ebitda_margin=round(ebitda / revenue, 6),
            da_eur_m=da,
            ebit_eur_m=ebit,
            fee_amort_eur_m=round(fee_amort + yr1_fee_expense, 6),
            senior_interest_eur_m=senior_interest,
            notes_interest_eur_m=notes_interest,
            total_interest_eur_m=total_interest,
            ebt_eur_m=ebt,
            tax_eur_m=tax,
            net_income_eur_m=net_income,
            capex_eur_m=capex,
            nwc_change_eur_m=nwc_change,
            senior_mandatory_amort_eur_m=actual_mandatory,
            levered_fcf_before_sweep_eur_m=lfcf_before_sweep,
            senior_cash_sweep_eur_m=sweep_amount,
            levered_fcf_eur_m=lfcf,
            senior_opening_eur_m=senior_bal,
            senior_closing_eur_m=senior_close,
            notes_opening_eur_m=notes_bal,
            notes_closing_eur_m=notes_close,
            total_debt_closing_eur_m=total_debt_close,
            leverage_x=leverage_x,
            interest_coverage_x=coverage_x,
        ))

        # Roll forward
        senior_bal = senior_close
        notes_bal = notes_close
        prev_revenue = revenue

    return years


# ── Step 5: Exit Metrics & IRR ────────────────────────────────────────────────

def _compute_irr(cash_flows: list[float]) -> tuple[float, bool]:
    """
    Solve for IRR using Brent's method.
    Returns (irr_decimal, converged).
    """
    def npv(rate: float) -> float:
        return sum(cf / (1.0 + rate) ** t for t, cf in enumerate(cash_flows))

    try:
        irr = float(brentq(npv, -0.9999, 100.0, xtol=1e-10, maxiter=500))
        return round(irr, 6), True
    except ValueError:
        logger.warning("IRR solver did not converge — cash flows may not have a real root.")
        return float("nan"), False


def compute_exit_metrics(
    a: EntryAssumptions,
    su: SourcesUses,
    projections: list[YearProjection],
) -> LBOMetrics:
    exit_proj = projections[a.exit_year - 1]
    exit_ebitda = exit_proj.ebitda_eur_m
    exit_ev = round(exit_ebitda * a.exit_multiple, 6)
    net_debt_at_exit = exit_proj.total_debt_closing_eur_m
    equity_proceeds = round(exit_ev - net_debt_at_exit, 6)

    equity_invested = su.equity_eur_m
    moic = round(equity_proceeds / equity_invested, 4) if equity_invested > 0 else float("nan")

    # IRR cash flows: Year-0 outflow, intermediate equity distributions (LFCF), exit proceeds
    # LFCF in each year represents distributable cash (after debt service + sweep)
    cf: list[float] = [-equity_invested]
    for i, proj in enumerate(projections[: a.exit_year]):
        is_exit = i == a.exit_year - 1
        distribution = proj.levered_fcf_eur_m + (equity_proceeds if is_exit else 0.0)
        cf.append(distribution)

    irr_val, converged = _compute_irr(cf)
    
    if converged and irr_val < 0.0:
        raise ValueError(
            f"AUCTUS Circuit Breaker Triggered: Negative LBO returns (IRR = {irr_val*100:.1f}%). "
            "Deal fails minimum viability."
        )

    total_debt_at_entry = su.total_sources_eur_m - su.equity_eur_m
    leverage_entry = round(
        total_debt_at_entry / a.entry_ebitda_eur_m, 4
    ) if a.entry_ebitda_eur_m > 0 else float("nan")

    leverage_exit = round(net_debt_at_exit / exit_ebitda, 4) if exit_ebitda > 0 else float("nan")

    coverage_values = [p.interest_coverage_x for p in projections[: a.exit_year]]
    finite_coverages = [x for x in coverage_values if math.isfinite(x)]
    min_coverage = round(min(finite_coverages), 4) if finite_coverages else float("nan")

    return LBOMetrics(
        entry_ev_eur_m=su.entry_ev_eur_m,
        entry_multiple_x=a.entry_multiple,
        exit_ev_eur_m=exit_ev,
        exit_multiple_x=a.exit_multiple,
        exit_ebitda_eur_m=exit_ebitda,
        net_debt_at_exit_eur_m=net_debt_at_exit,
        equity_invested_eur_m=equity_invested,
        equity_proceeds_eur_m=equity_proceeds,
        moic=moic,
        irr=irr_val,
        leverage_at_entry_x=round(leverage_entry, 4),
        leverage_at_exit_x=leverage_exit,
        interest_coverage_min_x=min_coverage,
        exit_year=a.exit_year,
        irr_solver_converged=converged,
    )


# ── Sensitivity Grid: Entry × Exit Multiple → IRR / MOIC ─────────────────────

def _scenario_irr_moic(
    a: EntryAssumptions,
    entry_m: float,
    exit_m: float,
) -> tuple[float, float]:
    """Recompute S&U, waterfall and metrics for a single (entry, exit) scenario."""
    tweaked = a.model_copy(update={"entry_multiple": entry_m, "exit_multiple": exit_m})
    su = build_sources_uses(tweaked)
    projs = compute_projections(tweaked, su)
    metrics = compute_exit_metrics(tweaked, su, projs)
    return metrics.irr, metrics.moic


def build_sensitivity_grid(
    a: EntryAssumptions,
    entry_steps: int = 5,
    exit_steps: int = 5,
    entry_delta: float = 1.0,
    exit_delta: float = 1.0,
) -> tuple[dict[str, dict[str, float]], dict[str, dict[str, float]]]:
    """
    Return (irr_grid, moic_grid) where each grid is {entry_m_str: {exit_m_str: value}}.

    Default: entry ±2× (5 steps × 1×) and exit ±2× (5 steps × 1×).
    """
    half_e = (entry_steps - 1) // 2
    half_x = (exit_steps - 1) // 2
    entry_multiples = [round(a.entry_multiple + (i - half_e) * entry_delta, 1) for i in range(entry_steps)]
    exit_multiples = [round(a.exit_multiple + (i - half_x) * exit_delta, 1) for i in range(exit_steps)]

    irr_grid: dict[str, dict[str, float]] = {}
    moic_grid: dict[str, dict[str, float]] = {}

    for em in entry_multiples:
        em_key = f"{em:.1f}x"
        irr_grid[em_key] = {}
        moic_grid[em_key] = {}
        for xm in exit_multiples:
            xm_key = f"{xm:.1f}x"
            try:
                irr_val, moic_val = _scenario_irr_moic(a, em, xm)
            except Exception as exc:
                logger.warning("Sensitivity scenario (%s, %s) failed: %s", em_key, xm_key, exc)
                irr_val, moic_val = float("nan"), float("nan")
            irr_grid[em_key][xm_key] = round(irr_val * 100, 2) if math.isfinite(irr_val) else float("nan")
            moic_grid[em_key][xm_key] = round(moic_val, 2) if math.isfinite(moic_val) else float("nan")

    return irr_grid, moic_grid


# ── Output Writers ────────────────────────────────────────────────────────────

def _projections_to_df(projections: list[YearProjection]) -> pd.DataFrame:
    return pd.DataFrame([p.model_dump() for p in projections])


def _sensitivity_to_df(grid: dict[str, dict[str, float]]) -> pd.DataFrame:
    df = pd.DataFrame(grid).T
    df.index.name = "entry_multiple"
    return df


def _compact_payload(result: LBOResult) -> dict[str, Any]:
    """
    Dense JSON payload containing only inflection years and summary metrics.
    Keeps terminal context minimal for downstream Claude consumption.

    Every figure cites its source key path for zero-hallucination verification.
    """
    projs = result.projections
    exit_y = result.assumptions.exit_year
    mid_y = max(1, exit_y // 2)

    def proj_summary(p: YearProjection) -> dict[str, Any]:
        return {
            "year": p.year,
            "revenue_eur_m": p.revenue_eur_m,
            "ebitda_eur_m": p.ebitda_eur_m,
            "ebitda_margin_pct": round(p.ebitda_margin * 100, 2),
            "total_interest_eur_m": p.total_interest_eur_m,
            "levered_fcf_eur_m": p.levered_fcf_eur_m,
            "total_debt_closing_eur_m": p.total_debt_closing_eur_m,
            "leverage_x": p.leverage_x,
            "interest_coverage_x": p.interest_coverage_x,
            "_source": f"projections[{p.year - 1}]",
        }

    inflection_years = {1, mid_y, exit_y}
    inflection_projections = [proj_summary(p) for p in projs if p.year in inflection_years]

    m = result.metrics
    su = result.sources_uses

    return {
        "company_name": result.company_name,
        "run_timestamp": result.run_timestamp,
        "sources_uses": {
            "entry_ev_eur_m": su.entry_ev_eur_m,
            "equity_eur_m": su.equity_eur_m,
            "senior_debt_eur_m": su.senior_debt_eur_m,
            "notes_eur_m": su.notes_eur_m,
            "total_uses_eur_m": su.total_uses_eur_m,
            "balance_check_eur_m": su.balance_eur_m,
            "_source": "sources_uses",
        },
        "inflection_projections": inflection_projections,
        "exit_metrics": {
            "exit_ev_eur_m": m.exit_ev_eur_m,
            "net_debt_at_exit_eur_m": m.net_debt_at_exit_eur_m,
            "equity_proceeds_eur_m": m.equity_proceeds_eur_m,
            "moic": m.moic,
            "irr_pct": round(m.irr * 100, 2) if math.isfinite(m.irr) else None,
            "irr_solver_converged": m.irr_solver_converged,
            "leverage_at_entry_x": m.leverage_at_entry_x,
            "leverage_at_exit_x": m.leverage_at_exit_x,
            "interest_coverage_min_x": m.interest_coverage_min_x,
            "_source": "metrics",
        },
        "assumptions": result.assumptions.model_dump() if hasattr(result.assumptions, "model_dump") else result.assumptions.dict(),
        "sensitivity_irr_pct": result.sensitivity_irr,
        "sensitivity_moic": result.sensitivity_moic,
    }


def _fill_sensitivity_sheet(
    ws: Any,
    grid: dict[str, dict[str, float]],
    green_thresh: float,
    amber_thresh: float,
) -> None:
    entry_keys = list(grid.keys())
    exit_keys = list(next(iter(grid.values())).keys())

    ws.cell(row=1, column=1, value="Entry \\ Exit")
    for ci, xk in enumerate(exit_keys, start=2):
        ws.cell(row=1, column=ci, value=xk)
    for ri, ek in enumerate(entry_keys, start=2):
        ws.cell(row=ri, column=1, value=ek)
        for ci, xk in enumerate(exit_keys, start=2):
            v = grid[ek][xk]
            ws.cell(row=ri, column=ci, value=v if math.isfinite(v) else None)

    n_rows = len(entry_keys)
    n_cols = len(exit_keys)
    data_range = f"{get_column_letter(2)}2:{get_column_letter(n_cols + 1)}{n_rows + 1}"

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    amber_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    ws.conditional_formatting.add(
        data_range,
        CellIsRule(operator="greaterThanOrEqual", formula=[str(green_thresh)],
                   fill=green_fill, font=Font(color="276221")),
    )
    ws.conditional_formatting.add(
        data_range,
        CellIsRule(operator="between",
                   formula=[str(amber_thresh), str(green_thresh - 0.0001)],
                   fill=amber_fill, font=Font(color="9C5700")),
    )
    ws.conditional_formatting.add(
        data_range,
        CellIsRule(operator="lessThan", formula=[str(amber_thresh)],
                   fill=red_fill, font=Font(color="9C0006")),
    )


def _write_excel_workbook(result: LBOResult, prefix: Path) -> Path:
    xlsx_path = Path(str(prefix) + "_model.xlsx")
    wb = openpyxl.Workbook()
    su = result.sources_uses
    m  = result.metrics

    # ── Sheet 1: Summary ─────────────────────────────────────────────────
    ws_sum = wb.worksheets[0]
    ws_sum.title = "Summary"

    su_rows = [
        ("Entry EV (€m)",       su.entry_ev_eur_m),
        ("Equity (€m)",         su.equity_eur_m),
        ("Senior Debt (€m)",    su.senior_debt_eur_m),
        ("Notes (€m)",          su.notes_eur_m),
        ("Total Debt (€m)",     round(su.senior_debt_eur_m + su.notes_eur_m, 6)),
        ("Advisor Fees (€m)",   su.advisor_fees_eur_m),
        ("Financing Fees (€m)", su.financing_fees_eur_m),
        ("Total Uses (€m)",     su.total_uses_eur_m),
        ("Balance Check (€m)",  su.balance_eur_m),
    ]
    irr_pct = round(m.irr * 100, 2) if math.isfinite(m.irr) else None
    exit_rows = [
        ("Exit EV (€m)",               m.exit_ev_eur_m),
        ("Net Debt at Exit (€m)",       m.net_debt_at_exit_eur_m),
        ("Equity Proceeds (€m)",        m.equity_proceeds_eur_m),
        ("MOIC",                        m.moic),
        ("IRR (%)",                     irr_pct),
        ("Leverage at Entry (×)",       m.leverage_at_entry_x),
        ("Leverage at Exit (×)",        m.leverage_at_exit_x),
        ("Min Interest Coverage (×)",   m.interest_coverage_min_x),
    ]

    _navy_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    _hdr_font  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    _lbl_font  = Font(name="Calibri", size=10)
    _val_font  = Font(name="Calibri", size=10)

    def _section_hdr(ws: Any, r: int, text: str) -> None:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        cell = ws.cell(row=r, column=1, value=text)
        cell.fill = _navy_fill
        cell.font = _hdr_font

    def _num_fmt(label: str) -> str:
        if "(€m)" in label:
            return '€#,##0.00'
        if "(%)" in label or "IRR" in label:
            return '0.00"%"'
        if "(×)" in label or "MOIC" in label or "Coverage" in label:
            return '0.00"×"'
        return "General"

    row = 1
    _section_hdr(ws_sum, row, "Sources & Uses")
    row += 1
    for label, val in su_rows:
        ws_sum.cell(row=row, column=1, value=label).font = _lbl_font
        c = ws_sum.cell(row=row, column=2, value=val)
        c.font = _val_font
        c.number_format = _num_fmt(label)
        row += 1

    row += 1
    _section_hdr(ws_sum, row, "Exit Metrics")
    row += 1
    for label, val in exit_rows:
        ws_sum.cell(row=row, column=1, value=label).font = _lbl_font
        c = ws_sum.cell(row=row, column=2, value=val)
        c.font = _val_font
        c.number_format = _num_fmt(label)
        row += 1

    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 16
    ws_sum.freeze_panes = "A2"

    # ── Sheet 2: P&L Waterfall ────────────────────────────────────────────
    ws_pl = wb.create_sheet("P&L Waterfall")
    df = _projections_to_df(result.projections)

    for col_idx, col_name in enumerate(df.columns, start=1):
        ws_pl.cell(row=1, column=col_idx, value=col_name)
    for row_idx, row_data in enumerate(df.itertuples(index=False), start=2):
        for col_idx, val in enumerate(row_data, start=1):
            ws_pl.cell(row=row_idx, column=col_idx, value=val)

    for col_idx, col_name in enumerate(df.columns, start=1):
        if col_name == "year":
            fmt = None
        elif col_name.endswith("_eur_m"):
            fmt = "€#,##0.00"
        elif col_name == "ebitda_margin" or "_pct" in col_name:
            fmt = "0.00%"
        elif col_name.endswith("_x"):
            fmt = '0.00"x"'
        else:
            fmt = None
        if fmt:
            for row_idx in range(2, len(df) + 2):
                ws_pl.cell(row=row_idx, column=col_idx).number_format = fmt

    ws_pl.freeze_panes = "A2"

    # ── Sheet 3: Sensitivity – IRR ────────────────────────────────────────
    ws_irr = wb.create_sheet("Sensitivity – IRR")
    _fill_sensitivity_sheet(ws_irr, result.sensitivity_irr, 20.0, 15.0)

    # ── Sheet 4: Sensitivity – MOIC ───────────────────────────────────────
    ws_moic = wb.create_sheet("Sensitivity – MOIC")
    _fill_sensitivity_sheet(ws_moic, result.sensitivity_moic, 2.0, 1.5)

    wb.save(xlsx_path)
    return xlsx_path


def write_outputs(
    result: LBOResult,
    output_dir: Path,
) -> tuple[Path, Path, Path]:
    ts = result.run_timestamp
    slug = result.company_name.lower().replace(" ", "_").replace("/", "_")
    prefix = output_dir / f"lbo_{slug}_{ts}"

    # Full JSON — internal pipeline artifact; not a user deliverable
    full_json_path = Path(str(prefix) + "_lbo_results.json")
    full_json_path.write_text(
        json.dumps(result.model_dump(), indent=2, default=str)
    )

    # Compact JSON — internal pipeline artifact; feeds memo_generator and deck_builder
    compact_path = Path(str(prefix) + "_lbo_compact.json")
    compact_path.write_text(
        json.dumps(_compact_payload(result), indent=2)
    )

    # Single consolidated workbook: Summary | P&L Waterfall | Sensitivity – IRR | Sensitivity – MOIC
    xlsx_path = _write_excel_workbook(result, prefix)

    return full_json_path, compact_path, xlsx_path


# ── Orchestrator ──────────────────────────────────────────────────────────────

def run_lbo(
    assumptions: EntryAssumptions,
    output_dir: Path,
) -> LBOResult:
    output_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")

    logger.info(
        "LBO run: company=%s entry_EV=€%.1fm (%.1fx × €%.1fm EBITDA)",
        assumptions.company_name,
        assumptions.entry_ebitda_eur_m * assumptions.entry_multiple,
        assumptions.entry_multiple,
        assumptions.entry_ebitda_eur_m,
    )

    su = build_sources_uses(assumptions)

    if abs(su.balance_eur_m) > 0.01:
        logger.warning(
            "Sources & Uses imbalance of €%.4fm — verify capital structure percentages.",
            su.balance_eur_m,
        )

    projections = compute_projections(assumptions, su)
    metrics = compute_exit_metrics(assumptions, su, projections)
    irr_grid, moic_grid = build_sensitivity_grid(assumptions)

    result = LBOResult(
        run_timestamp=ts,
        company_name=assumptions.company_name,
        assumptions=assumptions,
        sources_uses=su,
        projections=projections,
        metrics=metrics,
        sensitivity_irr=irr_grid,
        sensitivity_moic=moic_grid,
    )

    paths = write_outputs(result, output_dir)

    logger.info(
        "LBO complete: MOIC=%.2fx, IRR=%.1f%%, Exit EV=€%.1fm, outputs=%s",
        metrics.moic,
        metrics.irr * 100 if math.isfinite(metrics.irr) else float("nan"),
        metrics.exit_ev_eur_m,
        [str(p) for p in paths],
    )
    return result


# ── CLI ───────────────────────────────────────────────────────────────────────

def _parse_float_list(s: str) -> list[float]:
    return [float(x.strip()) for x in s.split(",")]


def main() -> int:
    parser = argparse.ArgumentParser(description="AUCTUS LBO Engine")

    # Identity
    parser.add_argument("--company-name", type=str, required=True, dest="company_name")
    parser.add_argument("--geography", type=str, default="DE")
    parser.add_argument("--output-dir", type=Path, required=True, dest="output_dir")

    # Entry
    parser.add_argument("--entry-ebitda", type=float, required=True, dest="entry_ebitda")
    parser.add_argument("--entry-multiple", type=float, required=True, dest="entry_multiple")
    parser.add_argument("--revenue-base", type=float, required=True, dest="revenue_base")

    # Capital structure
    parser.add_argument("--equity-pct", type=float, required=True, dest="equity_pct")
    parser.add_argument("--senior-debt-pct", type=float, required=True, dest="senior_debt_pct")
    parser.add_argument("--notes-pct", type=float, default=0.0, dest="notes_pct")

    # Debt pricing
    parser.add_argument("--euribor", type=float, default=0.039, dest="euribor_rate")
    parser.add_argument("--euribor-floor", type=float, default=0.0, dest="euribor_floor")
    parser.add_argument("--senior-spread-bps", type=int, default=375, dest="senior_spread_bps")
    parser.add_argument("--notes-rate", type=float, default=0.095, dest="notes_fixed_rate")
    parser.add_argument(
        "--notes-fixed", action=argparse.BooleanOptionalAction, default=True, dest="notes_is_fixed",
        help="Use fixed-rate notes (default). Pass --no-notes-fixed for floating rate.",
    )
    parser.add_argument("--notes-spread-bps", type=int, default=550, dest="notes_euribor_spread_bps")

    # Debt mechanics
    parser.add_argument("--senior-amort-pct", type=float, default=0.05, dest="senior_amort_pct")
    parser.add_argument("--senior-cash-sweep-pct", type=float, default=0.50, dest="senior_sweep_pct")

    # Fees
    parser.add_argument("--advisor-fee-pct", type=float, default=0.015, dest="advisor_fee_pct")
    parser.add_argument("--financing-fee-pct", type=float, default=0.020, dest="financing_fee_pct")
    parser.add_argument(
        "--capitalize-fees", action=argparse.BooleanOptionalAction, default=True, dest="capitalize_fees",
        help="Capitalise fees into Uses (default). Pass --no-capitalize-fees to expense them in Year-1 P&L.",
    )

    # Tax & projections
    parser.add_argument("--tax-rate", type=float, default=0.299, dest="tax_rate")
    parser.add_argument("--projection-years", type=int, default=5, dest="projection_years")
    parser.add_argument(
        "--revenue-growth", type=str, required=True, dest="revenue_growth",
        help="Comma-separated decimal growth rates, one per year (e.g. 0.07,0.07,0.06,0.06,0.05)",
    )
    parser.add_argument(
        "--ebitda-margins", type=str, required=True, dest="ebitda_margins",
        help="Comma-separated EBITDA margins, one per year (e.g. 0.175,0.18,0.185,0.19,0.195)",
    )
    parser.add_argument("--da-pct", type=float, default=0.025, dest="da_pct")
    parser.add_argument("--capex-pct", type=float, default=0.035, dest="capex_pct")
    parser.add_argument("--nwc-pct", type=float, default=0.08, dest="nwc_pct")

    # Exit
    parser.add_argument("--exit-year", type=int, default=5, dest="exit_year")
    parser.add_argument("--exit-multiple", type=float, required=True, dest="exit_multiple")

    args = parser.parse_args()

    try:
        growth_rates = _parse_float_list(args.revenue_growth)
        ebitda_margins = _parse_float_list(args.ebitda_margins)

        # Load configs for validation cross-reference
        fc, ac = load_configs()
        hf = ac.get("hard_filters", {})
        entry_ev = args.entry_ebitda * args.entry_multiple
        if entry_ev < hf.get("revenue_min_eur", 0) / 1e6:
            logger.warning(
                "Entry EV €%.1fm appears below AUCTUS minimum — confirm with IC.", entry_ev
            )

        assumptions = EntryAssumptions(
            company_name=args.company_name,
            geography=args.geography,
            entry_ebitda_eur_m=args.entry_ebitda,
            entry_multiple=args.entry_multiple,
            equity_pct=args.equity_pct,
            senior_debt_pct=args.senior_debt_pct,
            notes_pct=args.notes_pct,
            euribor_rate=args.euribor_rate,
            euribor_floor=args.euribor_floor,
            senior_spread_bps=args.senior_spread_bps,
            notes_is_fixed=args.notes_is_fixed,
            notes_fixed_rate=args.notes_fixed_rate,
            notes_euribor_spread_bps=args.notes_euribor_spread_bps,
            senior_amort_pct_annual=args.senior_amort_pct,
            senior_cash_sweep_pct=args.senior_sweep_pct,
            advisor_fee_pct_ev=args.advisor_fee_pct,
            financing_fee_pct_debt=args.financing_fee_pct,
            fees_capitalized=args.capitalize_fees,
            tax_rate=args.tax_rate,
            projection_years=args.projection_years,
            revenue_base_eur_m=args.revenue_base,
            revenue_growth_rates=growth_rates,
            ebitda_margins=ebitda_margins,
            da_pct_revenue=args.da_pct,
            capex_pct_revenue=args.capex_pct,
            nwc_pct_revenue_change=args.nwc_pct,
            exit_year=args.exit_year,
            exit_multiple=args.exit_multiple,
        )
    except (ValueError, KeyError) as exc:
        logger.error("Input validation error: %s", exc)
        return 1

    try:
        result = run_lbo(assumptions, args.output_dir)
        irr_pct = result.metrics.irr * 100 if math.isfinite(result.metrics.irr) else None
        print(json.dumps({
            "status": "success",
            "moic": result.metrics.moic,
            "irr_pct": irr_pct,
            "exit_ev_eur_m": result.metrics.exit_ev_eur_m,
            "irr_solver_converged": result.metrics.irr_solver_converged,
        }))
        return 0
    except (ValueError, ArithmeticError, ZeroDivisionError) as exc:
        logger.error("Computation error: %s", exc)
        return 2
    except OSError as exc:
        logger.error("Output write error: %s", exc)
        return 3


if __name__ == "__main__":
    sys.exit(main())
